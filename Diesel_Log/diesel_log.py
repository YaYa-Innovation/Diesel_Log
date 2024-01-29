from flask import Flask, render_template, request,jsonify
from openpyxl import *
import os

# Get the absolute path to the current directory
base_path = os.path.dirname(os.path.abspath(__file__))

# Use the absolute path for the Excel file
excel_file_path = os.path.join(base_path, 's.xlsx')
print(f"Absolute path to Excel file: {excel_file_path}")



app = Flask(__name__)

@app.route('/')
def index():
    return render_template('index.html', current_date='2024-01-29')

@app.route('/submit', methods=['POST'])
def submit():
    # Handle form submission logic here
    return 'Form submitted successfully!'

@app.route('/login', methods=['POST'])
def login():
    # Retrieve data from the login form
    username = request.form.get('username')
    password = request.form.get('password')

    # Perform any necessary validation or logic here

    # Open the Excel workbook and get the active sheet
    workbook =load_workbook(excel_file_path)  # Provide the correct path to your Excel file
    sheet = workbook.active

    # Append the login data to the Excel sheet
    sheet.append([username, password])

    # Save the changes to the Excel file
    workbook.save(excel_file_path)  # Provide the correct path to your Excel file

    return 'Login successful! Data saved to Excel.'
def get_age_by_username(username):
    # Open the Excel workbook and get the active sheet
    workbook = load_workbook(excel_file_path)
    sheet = workbook.active

    # Iterate through rows to find a matching username
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == username:  # Assuming username is in the first column (adjust as needed)
            return str(row[1])  # Assuming age is in the second column (adjust as needed)

    return ''  # Return empty string if no matching username is found

@app.route('/check_username', methods=['POST'])
def check_username():
    # Retrieve the username from the JSON request
    username = request.json.get('username')

    # Perform a check to see if a matching age exists for the provided username
    matching_age = get_age_by_username(username)

    # Return the matching age (or an empty string if no match)
    return jsonify({'age': matching_age, 'username': username})
if __name__ == '__main__':
    app.run(debug=True)
